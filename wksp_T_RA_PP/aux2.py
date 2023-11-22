import os

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

project_pth: str = os.getcwd()
res_pth: str = project_pth + "\\res"
destiny_pth: str = project_pth + "\\destiny_files"
origin_pth: str = project_pth + "\\origin_files"

lst = os.listdir(origin_pth)
out = open(destiny_pth + "\\out.txt", "w")
for i in lst:
    tmp: Workbook = openpyxl.load_workbook(origin_pth + "\\" + i)
    shts: list[Worksheet] = tmp.worksheets
    for sh in shts:
        if "=" not in sh.title:
            continue
        for row in range(6, sh.max_row + 1):
            desde_bornera = sh.cell(row, 6).value
            if desde_bornera is None or desde_bornera == "":
                continue
            tag = sh.cell(row, 2).value.strip("-")
            desde = sh.cell(row, 5).value.strip("=")
            desde_borne = sh.cell(row, 7).value

            hasta = sh.cell(row, 8).value.strip("=")
            hasta_bornera = sh.cell(row, 9).value
            hasta_borne = sh.cell(row, 10).value

            out.write(tag + "\t")
            out.write(str(desde) + "\t")
            out.write(str(desde_bornera) + "\t")
            out.write(str(desde_borne) + "\t")
            out.write(str(hasta) + "\t")
            out.write(str(hasta_bornera) + "\t")
            out.write(str(hasta_borne) + "\t")
            out.write("\n")
            print("terminada fila " + str(row))
        pass


out.close()
