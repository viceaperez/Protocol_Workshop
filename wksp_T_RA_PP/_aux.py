import os

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

project_pth: str = os.getcwd()
res_pth: str = project_pth + "\\res"
destiny_pth: str = project_pth + "\\destiny_files"
origin_pth: str = project_pth + "\\origin_files"

lst = os.listdir(origin_pth)

lst_log: dict[str:str] = {}

out = open(destiny_pth + "\\out.txt", "w")
for e in lst:

    '''
    tmp: Workbook = openpyxl.load_workbook(origin_pth + "\\" + e)
    sh: Worksheet = tmp.worksheets[0]
    tag = sh.cell(19, 9).value
    corr = sh.cell(12, 23).value
    '''

    corr = e.split("-")[0]
    tag = e.split("_")[1].strip(".xlsx")
    out.write(tag + "\t" + str(corr) + "\n")
    pass

out.close()
