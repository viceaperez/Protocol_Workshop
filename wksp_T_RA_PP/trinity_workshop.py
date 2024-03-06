import os
import random
import re

import numpy
import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class TrinityWorkshop:
    project_pth: str = os.getcwd()
    res_pth: str = project_pth + "\\res"
    destiny_pth: str = project_pth + "\\destiny_files"
    origin_pth: str = project_pth + "\\origin_files"
    src_pth = res_pth + "\\MATRIZ DE TENDIDO.txt"
    wb = open(src_pth)

    starting_corr = 0
    db: list[dict[str:Cell]] = []

    @classmethod
    def fetch(cls):
        for row in cls.wb:
            pts = row.split("\t")
            tipo_cable = pts[1].strip()
            fields: dict[str:str] = {
                "tag": pts[0].strip(),
                "tipo_cable": tipo_cable,
                "desde": pts[2].strip(),
                "hasta": pts[3].strip(),
                "n_hebras": cls.get_hebras(None, tipo_cable.split(" ")[1], None),
                "ubicacion": pts[5].strip(),
                "largo": pts[4].strip(),
                "uso": cls.resolve_uso(pts[0].strip())
            }
            cls.db.append(fields)
            pass
        pass

    @classmethod
    def find(cls, tag):
        for f in cls.db:
            if f["tag"] == tag:
                return f
            pass
        pass

    @classmethod
    def elem(cls, idx):
        return cls.db[idx]

    pass

    @classmethod
    def len(cls):
        return len(cls.db)

    planos: dict[str:str] = {
        "Sala3_elem": "SNN4008-E-MMI-10-CP-PL-0004-L0001",
        "Sala4_elem": "SNN4008-E-MMI-10-CP-PL-0006-L0001",
        "Sala5_elem": "SNN4008-E-MMI-10-CP-PL-0008-L0001",
        "Sala6_elem": "SNN4008-E-MMI-10-CP-PL-0010-L0001",
        "SSGG_elem": "SNN4008-E-MMI-10-CP-PL-0002-L0001",
        "Sala3_ilum": "SNN4008-E-MMI-15-EL-PL-0005-L0001",
        "Sala4_ilum": "SNN4008-E-MMI-15-EL-PL-0005-L0001",
        "Sala5_ilum": "SNN4008-E-MMI-15-EL-PL-0005-L0001",
        "Sala6_ilum": "SNN4008-E-MMI-15-EL-PL-0005-L0001",
        "SSGG_ilum": "SNN4008-E-MMI-15-EL-PL-0004-L0001",
    }

    @classmethod
    def resolve_plano(cls, ub, desde, hasta):
        if ub:
            if re.match("220", ub):
                return cls.planos["PJ"]
            if re.match("casa", ub, re.IGNORECASE):
                return cls.planos["SSGG"]
            pass
        desde_main = desde.split("+")[0]
        hasta_main = hasta.split("+")[0]
        if re.match("SJD1", desde_main):
            if re.match("SJD1", hasta_main):
                return cls.planos["SalaPJ_1_2"]
            if re.match("SJD2", hasta_main):
                return cls.planos["SalaPJ_1_2"]
            if re.match("SJD3", hasta_main):
                return cls.planos["PJ"]

        if re.match("^tdc(.)+$", desde_main, re.IGNORECASE):
            return cls.planos["SSGG"]
        if re.match("^tgc(.)+$", desde_main, re.IGNORECASE):
            return cls.planos["SSGG"]
        if re.match("^tdc(.)+$", hasta_main, re.IGNORECASE):
            return cls.planos["SSGG"]
        if re.match("^tgc(.)+$", hasta_main, re.IGNORECASE):
            return cls.planos["SSGG"]
        if re.match("^PK(.)+$", desde_main, re.IGNORECASE):
            return cls.planos["PK"]
        if re.match("^PJ(.)+$", desde_main, re.IGNORECASE):
            return cls.planos["PJ"]

        if re.match("PZ(.)+", desde_main):
            if re.match("SKD", hasta_main):
                return cls.planos["Planta"]
            if re.match("PCZ(.)+", hasta):
                return cls.planos["PZ"]
        if re.match("PATR(.)+", desde):
            if re.match("SKD(.)+", hasta):
                return cls.planos["Planta"]
            if re.match("PATR(.)+", hasta):
                return cls.planos["PATR"]
        if re.match("(.)*(BAT)(.)*", desde):
            if re.match("(.)*(BAT)(.)*", hasta):
                return cls.planos["SSGG"]
        #raise Exception("A: " + desde + "\nB: " + hasta + "\nNo coincide con plano")
        # print("A: " + desde + "\nB: " + hasta + "\nNo coincide con plano")

        pass

    @classmethod
    def resolve_uso(cls, tag):
        if re.match("W(.)+$", tag, re.IGNORECASE):
            return "C"
        return "F"
        pass

    @classmethod
    def get_hebras(cls, n_hebras, calibre, homologacion):
        if homologacion:
            tmp = homologacion.split(" ")[0]
            tmp2 = re.sub(re.compile("([^0-9])*(X)(.)*", re.IGNORECASE), "", tmp)
            return int(tmp2)
        if n_hebras:
            return int(n_hebras)
        if "c" in calibre:
            return int(calibre.strip().split("-c-")[0])
        else:
            return int(calibre.strip().split("x")[0])

    pass

    @classmethod
    def ensure_paths(cls):
        try:
            os.makedirs(cls.res_pth, exist_ok=True)
            os.makedirs(cls.origin_pth, exist_ok=True)
            os.makedirs(cls.destiny_pth, exist_ok=True)
        except OSError:
            pass
        pass


class TendidoWorkshop:
    res_pth = TrinityWorkshop.res_pth + "\\T"
    working_template: Workbook = openpyxl.load_workbook(
        res_pth + "\\SNN4008-E-MMI-01-ELE-006 Protocolo Tendido de conductores electricos AT.xlsx")
    working_ws: Worksheet = working_template.worksheets[0]

    tags_to_gen = []
    corrs = []
    fl = open(TrinityWorkshop.res_pth + "\\out.txt")
    for line in fl:
        pts = line.strip().split("\t")
        tags_to_gen.append(pts[0])
        corrs.append(pts[1])
        pass

    @classmethod
    def set(cls, r, c, val):
        cls.working_ws.cell(r, c, val)
        pass

    fields: dict[str:Cell] = {
        "subestacion": working_ws.cell(10, 9),
        "correlativo": working_ws.cell(12, 23),
        "plano": working_ws.cell(13, 7),
        "fecha": working_ws.cell(13, 23),
        "check_control": working_ws.cell(17, 7),
        "check_alumbrado": working_ws.cell(17, 14),
        "check_fuerza": working_ws.cell(17, 20),
        "check_pantalla_y": working_ws.cell(17, 29),
        "check_pantalla_n": working_ws.cell(17, 33),
        "tag": working_ws.cell(19, 9),
        "seccion": working_ws.cell(20, 9),
        "aislacion": working_ws.cell(21, 9),
        "desde": working_ws.cell(19, 23),
        "hasta": working_ws.cell(20, 23),
        "longitud": working_ws.cell(21, 23),
        "elabora_nombre": working_ws.cell(44, 4),
        "elabora_cargo": working_ws.cell(45, 4),
        "elabora_fecha": working_ws.cell(46, 4),
        "revisa_nombre": working_ws.cell(44, 12),
        "revisa_cargo": working_ws.cell(45, 12),
        "revisa_fecha": working_ws.cell(46, 12),
    }

    @classmethod
    def fetch(cls):
        pass

    @classmethod
    def inprint(cls, fields: dict[str:Cell], corr):

        cls.fields["correlativo"].value = corr
        cls.fields["plano"].value = TrinityWorkshop.resolve_plano(fields["ubicacion"], fields["tag"])
        cls.fields["fecha"].value = "25/02/2024"
        cls.toggle_field(fields["tag"])
        cls.fields["tag"].value = fields["tag"]
        cls.fields["desde"].value = fields["desde"]
        cls.fields["hasta"].value = fields["hasta"]
        cls.fields["longitud"].value = fields["largo"]
        cls.toggle_checks()
        cls.fields["elabora_nombre"].value = "Jose Godoy"
        cls.fields["elabora_cargo"].value = "Supervisor"
        # cls.fields["elabora_fecha"].value=corr
        cls.fields["revisa_nombre"].value = "Claudio Boris Hurtado G."
        cls.fields["revisa_cargo"].value = "Jefe Terreno"
        # cls.fields["revisa_fecha"].value=corr

        pts = fields["tipo_cable"].split(" ")
        cls.fields["seccion"].value = pts[1] + " " + pts[2]
        cls.fields["aislacion"].value = pts[0]
        pass

    @classmethod
    def gen(cls):
        for i in range(len(cls.tags_to_gen)):
            fields = TrinityWorkshop.find(cls.tags_to_gen[i])
            cls.inprint(fields, cls.corrs[i])
            cls.working_template.save(
                TrinityWorkshop.destiny_pth + "\\" + cls.corrs[i] + "-T_" + fields["tag"] + ".xlsx")
        pass

    pass

    @classmethod
    def toggle_field(cls, tag):
        if re.match("W(.)+$", tag, re.IGNORECASE):
            cls.fields["check_control"].value = "✔"
            cls.fields["check_pantalla_y"].value = "✔"
            cls.fields["check_fuerza"].value = ""
            cls.fields["check_pantalla_n"].value = ""
            return
        cls.fields["check_fuerza"].value = "✔"
        cls.fields["check_pantalla_n"].value = "✔"
        cls.fields["check_control"].value = ""
        cls.fields["check_pantalla_y"].value = ""
        pass

        pass

    @classmethod
    def toggle_checks(cls):
        for i in range(27, 36):
            cls.working_ws.cell(i, 16, "✔")
        pass

    pass


class AislacionWorkshop:
    res_pth = TrinityWorkshop.res_pth + "\\RA"
    working_template: Workbook = openpyxl.load_workbook(res_pth + "\\SNN4008-E-MMI-01-ELE-004 Pruebas de aislación de cables REV.0 (1).xlsx")
    working_ws: Worksheet = working_template.worksheets[0]

    tags_to_gen = []
    corrs = []
    fl = open(TrinityWorkshop.res_pth + "\\out.txt")
    for line in fl:
        pts = line.strip().split("\t")
        tags_to_gen.append(pts[0])
        corrs.append(pts[1])
        pass

    fields: dict[str: Cell] = {
        "plano": working_ws.cell(11, 3),
        "correlativo": working_ws.cell(13, 3),
        "fecha": working_ws.cell(15, 3),
        "check_control": working_ws.cell(20, 2),
        "check_alumbrado": working_ws.cell(22, 2),
        "check_fuerza": working_ws.cell(24, 2),
        "check_pantalla_y": working_ws.cell(21, 5),
        "check_pantalla_n": working_ws.cell(23, 5),
        "tag": working_ws.cell(26, 3),
        "seccion": working_ws.cell(27, 3),
        "aislacion": working_ws.cell(28, 3),
        "nivel_tension": working_ws.cell(10, 3),
        "tension_serv": working_ws.cell(29, 3),
        "desde": working_ws.cell(30, 3),
        "hasta": working_ws.cell(31, 3),
        "longitud": working_ws.cell(32, 3),
        "capacidad": working_ws.cell(33, 3),
        "instrumento_tipo": working_ws.cell(36, 3),
        "instrumento_marca": working_ws.cell(37, 3),
        "instrumento_modelo": working_ws.cell(38, 3),
        "instrumento_serie": working_ws.cell(39, 3),
        "instrumento_calibracion": working_ws.cell(40, 3),
        "ensayo_tension": working_ws.cell(43, 3),
        "ensayo_tiempo": working_ws.cell(44, 3),
        "ensayo_temperatura": working_ws.cell(45, 3),
        "ensayo_humedad": working_ws.cell(46, 3),
        "ensayo_longitud": working_ws.cell(47, 3),
        "elabora_nombre": working_ws.cell(51, 2),
        "elabora_cargo": working_ws.cell(52, 2),
        "elabora_fecha": working_ws.cell(53, 2),
        "revisa_nombre": working_ws.cell(51, 7),
        "revisa_cargo": working_ws.cell(52, 7),
        "revisa_fecha": working_ws.cell(53, 7)
    }

    capacidades = None

    @classmethod
    def inprint(cls, fields: dict[str:Cell], corr):

        #cls.fields["plano"].value = TrinityWorkshop.resolve_plano(fields["ubicacion"], fields["desde"], fields["hasta"])
        cls.fields["correlativo"].value = corr
        cls.fields["fecha"].value = "26/02/2024"
        cls.toggle_checks(fields["tag"])
        cls.fields["tag"].value = fields["tag"]
        cls.fields["nivel_tension"].value = "220-400 V"
        cls.fields["tension_serv"].value = "220-400 V"
        cls.fields["desde"].value = fields["desde"]
        cls.fields["hasta"].value = fields["hasta"]
        cls.fields["longitud"].value = fields["largo"]
        cls.fields["instrumento_tipo"].value = "Multímetro"
        cls.fields["instrumento_marca"].value = "Fluke"
        cls.fields["instrumento_modelo"].value = "1507"
        cls.fields["instrumento_serie"].value = "25443"
        cls.fields["instrumento_calibracion"].value = ""
        cls.fields["ensayo_tension"].value = ""
        cls.fields["ensayo_tiempo"].value = ""
        cls.fields["ensayo_temperatura"].value = ""
        cls.fields["ensayo_humedad"].value = ""
        cls.fields["ensayo_longitud"].value = ""
        cls.fields["elabora_nombre"].value = "Jose Godoy Espinoza"
        cls.fields["elabora_cargo"].value = "Supervisor Eléctrico"
        cls.fields["elabora_fecha"].value = ""
        cls.fields["revisa_nombre"].value = "Claudio Boris H."
        cls.fields["revisa_cargo"].value = "Jefe Terreno"
        cls.fields["revisa_fecha"].value = ""

        pts = fields["tipo_cable"].split(" ")
        cls.fields["capacidad"].value = cls.resolve_capacidad(pts[1], pts[2]) + " A"
        cls.fields["seccion"].value = pts[1] + " " + pts[2]
        cls.fields["aislacion"].value = pts[0]

        cls.populate_table(fields["n_hebras"], pts[1], fields["uso"])

        pass

    pass

    @classmethod
    def fetch(cls):
        cls.capacidades = []
        tmp = open(cls.res_pth + "\\tabla corrientes.txt")
        for l in tmp:
            t = l.split(" ")
            cls.capacidades.append([t[0], t[1], t[2]])
            pass
        pass

    @classmethod
    def toggle_checks(cls, tag):
        if re.match("W(.)+$", tag, re.IGNORECASE):
            cls.fields["check_control"].value = "✔"
            cls.fields["check_pantalla_y"].value = "✔"
            cls.fields["check_fuerza"].value = ""
            cls.fields["check_pantalla_n"].value = ""
            return
        cls.fields["check_fuerza"].value = "✔"
        cls.fields["check_pantalla_n"].value = "✔"
        cls.fields["check_control"].value = ""
        cls.fields["check_pantalla_y"].value = ""
        pass

    @classmethod
    def resolve_capacidad(cls, calibre: str, unidad: str):
        offset = 0
        if re.match("AWG", unidad, re.IGNORECASE):
            offset += 1
        seccion = cls.get_seccion(calibre)
        for e in cls.capacidades:
            if e[offset] == seccion:
                return e[2]
        raise Exception("calibre no encontrado")
        pass

    @classmethod
    def gen_temp(cls):
        num = random.Random().randint(a=150, b=250)
        return str(num / 10)
        pass

    @classmethod
    def gen_hum(cls):
        num = random.Random().randint(a=300, b=500)
        return str(num / 10)
        pass

    @classmethod
    def gen_res(cls):
        num = numpy.random.normal(loc=100, scale=15)
        return str(int(num) / 10)

    @classmethod
    def get_hebras(cls, calibre):
        if "c" in calibre:
            return calibre.strip().split("-c-")[0].replace("-", "/")
        else:
            return calibre.strip().split("x")[0].replace("-", "/")
        pass

    @classmethod
    def get_seccion(cls, calibre):
        if "c" in calibre:
            pts = calibre.strip().split("-c-")
            return pts[1].replace("-", "/")
        else:
            pts = calibre.strip().split("x")
            return pts[len(pts) - 1].replace("-", "/")
        pass

    @classmethod
    def populate_table(cls, n_hebras, calibre, uso):
        if n_hebras:
            hebras = int(n_hebras)
        else:
            hebras = int(cls.get_hebras(calibre))
            pass

        for i in range(0, hebras):
            row = 12 + (2 * i)
            for j in range(i, hebras):
                if i == j:
                    continue
                col = 8 + (2 * j)
                #cls.working_ws.cell(row, col).value = cls.gen_res()
                pass
            if uso == "C":
                #cls.working_ws.cell(row, 40).value = cls.gen_res()
                #cls.working_ws.cell(row, 41).value = cls.gen_res()
                pass
            cls.working_ws.cell(row, 42).value = "OK"

        pass

    @classmethod
    def gen(cls):
        for i in range(len(cls.tags_to_gen)):
            cls.flush()
            fields = TrinityWorkshop.find(cls.tags_to_gen[i])

            if fields["n_hebras"] != "":
                if int(fields["n_hebras"]) <= 1:
                    continue
            elif int(cls.get_hebras(fields["tipo_cable"].split(" ")[0])) <= 1:
                continue

            cls.inprint(fields, cls.corrs[i])
            cls.working_template.save(
                TrinityWorkshop.destiny_pth + "\\" + cls.corrs[i] + "-RA_" + fields["tag"] + ".xlsx")
        pass

    @classmethod
    def flush(cls):
        for i in range(0, 16):
            row = 12 + (2 * i)
            for j in range(0, 16):
                if i == j:
                    continue
                col = 8 + (2 * j)
                cls.working_ws.cell(row, col).value = ""
                pass

            cls.working_ws.cell(row, 40).value = ""
            cls.working_ws.cell(row, 41).value = ""
            cls.working_ws.cell(row, 42).value = ""
        pass


kind = [
 #   "T",
    "RA",
]


def start(kinds: list[str]):
    TrinityWorkshop.ensure_paths()
    TrinityWorkshop.fetch()
    if "T" in kinds:
        TendidoWorkshop.gen()
    if "RA" in kinds:
        AislacionWorkshop.fetch()
        AislacionWorkshop.gen()
    pass


start(kind)

pass
