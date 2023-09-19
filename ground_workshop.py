import os

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

project_pth: str = os.getcwd()
res_pth: str = project_pth + "\\res"
destiny_pth: str = project_pth + "\\destiny_files"
base_fl_pth: str = res_pth + "\\base.xlsx"
base_fl: Workbook = openpyxl.load_workbook(base_fl_pth)
lista_fl_pth: str = res_pth + "\\lista.xlsx"
lista_fl: Workbook = openpyxl.load_workbook(lista_fl_pth, read_only=True)

sectores = {
    "PK": "Patio 500kV",
    "PJ": "Patio 220kV",
    "PATR": "Patio ATR",
    "PZ": "Patio Reactores"
}

elementos = {
    "malla": "Puesta a Tierra Subterranea",
    "cerco": "Cerco",
}

planos = {
    "planta": "SNN4008-E-PRN-13-EL-PL-0001-L0001",
    "pk": "SNN4008-E-PRN-13-EL-PL-0001-L0002",
    "pk-d": "SNN4008-E-PRN-13-EL-PL-0001-L0003",
    "pj": "SNN4008-E-PRN-13-EL-PL-0006-L0001",
    "pj-d": "SNN4008-E-PRN-13-EL-PL-0006-L0002",
    "atr_r": "SNN4008-E-PRN-13-EL-PL-0007-L0001",
    "atr_r-d": "SNN4008-E-PRN-13-EL-PL-0007-L0002",
}

ws: Worksheet = base_fl.worksheets[0]
ws1: Worksheet = lista_fl.worksheets[0]

checks = {
    "Bancoductos": ws.cell(12, 12),
    "Estructurales": ws.cell(12, 18),
    "Escalerillas": ws.cell(12, 24),
    "Camaras": ws.cell(14, 12),
    "Equipos": ws.cell(14, 18),
    "totros": ws.cell(14, 24),
}


def set(r, c, val):
    ws.cell(r, c).value = val
    pass


def get(r, c):
    val = str(ws1.cell(r, c).value)
    if val == "None":
        return "-"
    if val == "0":
        return "-"
    return val


max_line = ws1.max_row


def flush_base():
    for check in checks.values():
        check.value = ""
        pass

    for j in range(0, 3):
        set(19, 9 + (2 * j), "-")
        pass

    for j in range(0, 3):
        set(20, 9 + (2 * j), "-")
        pass

    for j in range(0, 5):
        for k in range(0, 2):
            set(18 + j, 22 + (k * 7), "-")
            pass
        pass
    pass


def resolve_elemento(tag_elem):  # todo refino de otros elementos
    if "Diagonal" in tag_elem:
        return elementos["malla"]
    if "Cerco" in tag_elem:
        return elementos["cerco"]

    if "Patio" in tag_elem:
        return elementos["malla"]
    pass


def resolve_plano(elemento, sector):  # todo euristica get plano
    if sector == "PK":
        if elemento == elementos["malla"]:
            return planos["pk"]

        return planos["pk"]
    if sector == "PJ":
        if elemento == elementos["malla"]:
            return planos["pj"]
        return planos["pj"]
    if sector == "PATR" or sector == "PZ":
        if elemento == elementos["malla"]:
            return planos["atr_r"]
        return planos["atr_r"]
    pass


def toggle_check(elemento):  # todo
    if elemento == elementos["malla"]:
        checks["totros"].value = "✔"
        return
    if elemento == elementos["cerco"]:
        checks["Estructurales"].value = "✔"
        checks["totros"].value = "✔"
    pass


for i in range(2, max_line):
    estado_interno = get(i, 5)
    if estado_interno == "Entregado a calidad":
        continue
        pass

    if estado_interno != "go":
        continue
        pass

    flush_base()

    corr: str = get(i, 1).zfill(3)
    set(6, 29, corr)

    sector = get(i, 4)
    set(8, 5, sectores[sector])

    tag = get(i, 2)
    set(8, 24, tag)

    elemento: str = resolve_elemento(tag)
    set(7, 6, elemento)

    plano: str = resolve_plano(elemento, sector)
    set(9, 9, plano)

    toggle_check(elemento)

    for j in range(0, 3):
        cable_4 = get(i, 6 + j)
        print(cable_4)
        if cable_4 == "-":
            set(19, 9 + (2 * j), cable_4)
        else:
            set(19, 9 + (2 * j), float(cable_4))

        pass

    for j in range(0, 3):
        cable_2 = get(i, 9 + j)
        if cable_2 == "-":
            set(20, 9 + (2 * j), cable_2)
        else:
            set(20, 9 + (2 * j), float(cable_2))
        pass

    for j in range(0, 5):
        for k in range(0, 2):
            cant_sold: str
            cant_sold = get(i, 12 + (2 * j) + k)
            set(18 + j, 22 + (k * 7), cant_sold)
            pass
        pass

    for j in range(0,12):
        set(25+j,15,"✔")

    set(19, 15, "=SUM(I19:N19)")
    set(20, 15, "=SUM(I20:N20)")

    set(49, 5, "Camilo Miño")
    set(50, 5, "Supervisor Eléctrico")
    set(54, 5, "")  # TODO fecha de los protocolos

    set(49, 12, "Claudio Boris Hurtado G.")
    set(50, 12, "Jefe Terreno")
    # set(54,12,"") #TODO fecha de los protocolos

    destiny_fl: str = destiny_pth + "\\" + corr + "-PMT.xlsx"
    base_fl.save(destiny_fl)

    pass

pass
