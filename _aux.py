import os

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ot = open("log.txt", "w")

project_pth: str = os.getcwd()
res_pth: str = project_pth + "\\SSGG"

lst = os.listdir(res_pth)

for fl in lst:
    working_template: Workbook = openpyxl.load_workbook(res_pth + "\\" + fl, read_only=True)
    working_ws: Worksheet = working_template.worksheets[0]
    corr = working_ws.cell(12, 23).value
    tag = working_ws.cell(19, 9).value
    desde = working_ws.cell(19, 23).value
    hasta = working_ws.cell(20, 23).value
    print(tag + "\t" + desde + "\t" + hasta + "\n")
    ot.write(str(corr) + "\t" + tag + "\t" + desde + "\t" + hasta + "\n")
    working_template: Workbook = openpyxl.load_workbook(project_pth + "\\pp\\" + str(corr) + "-PP_" + tag + ".xlsx",
                                                        read_only=True)
    working_ws: Worksheet = working_template.worksheets[0]
    row = 28
    col = 2
    regleta_or = working_ws.cell(row, col).value
    while regleta_or:
        borne_or = working_ws.cell(row, col + 2).value
        hebra_or = working_ws.cell(row, col + 4).value

        regleta_des = working_ws.cell(row, 34).value
        borne_des = working_ws.cell(row, 32).value
        hebra_des = working_ws.cell(row, 30).value

        ot.write("\t"+regleta_or + ":" + borne_or + "\t" + regleta_des + ":" + borne_des + "\n")

        row += 1
        regleta_or = working_ws.cell(row, col).value
        pass

    pass
ot.close()
