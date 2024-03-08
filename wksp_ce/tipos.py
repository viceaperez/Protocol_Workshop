import os
import re

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# TODO: MODIFICAR EL ORIGEN DE DATOS DE ELEMENTOS PARA DISTINTOS PATIOS
origen = "salas.xlsx"

project_pth: str = os.getcwd()
res_pth: str = project_pth + "\\res"


class Elemento:

    def __init__(self):
        self.tipo = None
        self.boxes: list[float] = []
        self.nombre = ""
        self.patron: re.Pattern
        self.area_exp: re.Pattern

    pass


elementos: dict[str:Elemento] = {}
''''''
wb: Workbook = openpyxl.load_workbook(res_pth + "\\" + origen)
for ws in wb:
    ws: Worksheet
    el = Elemento()
    el.nombre = ws.cell(24, 2).value
    el.id = ws.title
    el.patron = re.compile(ws.cell(22, 2).value, re.IGNORECASE)
    el.area_exp = re.compile(ws.cell(23, 2).value, re.IGNORECASE)
    for i in range(1, 21):
        el.boxes.append(ws.cell(i, 2).value)
        pass
    elementos[el.id] = el
    pass
