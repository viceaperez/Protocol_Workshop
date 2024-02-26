import os

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from wksp_ce.tipos import Elemento, elementos

project_pth: str = os.getcwd()
res_pth: str = project_pth + "\\res"
destiny_pth: str = project_pth + "\\destiny_files"
base_fl_pth: str = res_pth + "\\SNN4008-E-MMI-01-ELE-003 Canalización Eléctrica Rev.0.xlsx"
base_fl: Workbook = openpyxl.load_workbook(base_fl_pth)
lista_fl_pth: str = res_pth + "\\LOG Protocolos Canalizado MMI.xlsx"
lista_fl: Workbook = openpyxl.load_workbook(lista_fl_pth, read_only=True)


def ensure_paths():
    lst: list[str] = ["res", "destiny_files"]
    dir_tree: list[str] = os.listdir(project_pth)
    for pth in lst:
        if pth not in dir_tree:
            os.mkdir(pth)
        pass
    pass


ensure_paths()

sectores = {
    "SJD3": "Sala 3",
    "SJD4": "Sala 4",
    "SJD5": "Sala 5",
    "SJD6": "Sala 6",
    "SSGG": "Sala Servicios Generales",
}

planos = {
    "SJD3": "SNN4008-E-MMI-14-EL-PL-0003-L0001",
    "SJD4": "SNN4008-E-MMI-14-EL-PL-0003-L0001",
    "SJD5": "SNN4008-E-MMI-14-EL-PL-0004-L0001",
    "SJD6": "SNN4008-E-MMI-14-EL-PL-0004-L0001",
    "SSGG": "SNN4008-E-MMI-14-EL-PL-0002-L0001",
}

ws: Worksheet = base_fl.worksheets[0]
ws1: Worksheet = lista_fl.worksheets[0]

fields: dict[str:Cell] = {
    "Area": ws.cell(11, 6),
    "Elemento": ws.cell(11, 26),
    "Correlativo": ws.cell(8, 59),
    "Plano": ws.cell(12, 7),
}

boxes: list[Cell] = []
for i in range(0, 3):
    for j in range(0, 6):
        boxes.append(ws.cell(15 + j, (6 * (i * i)) + (16 * i) + 22))
        # 22, 44, 60
    pass


def set(r, c, val):
    ws.cell(r, c).value = val
    pass


def get(r, c):
    val = str(ws1.cell(r, c).value)
    if val == "None":
        return ""
    if val == "0":
        return ""
    return val


max_line = ws1.max_row


def flush_base():
    for box in boxes:
        box.value = ""
        set(box.row, box.column - 3, "")
        pass


def resolve_elemento(tag_elem, area_elem) -> Elemento:
    for tipo in elementos.values():
        if tipo.patron.match(tag_elem) and tipo.area_exp.match(area_elem):
            return tipo
    pass


def resolve_plano(sector):
    return planos[sector]


def toggle_boxes(elemento: Elemento):
    for i in range(len(boxes)):
        box = boxes[i]
        val = elemento.boxes[i]
        box.value = val
        if val and val > 0:
            set(box.row, box.column-3, "✔")
        pass


def toggle_approval():
    for j in range(0, 11):
        set(26 + j, 38, "✔")
    pass


def toggle_signatures():
    # TODO: ACTUALIZAR NOMBE SUPERVISOR
    nombre_elabora = "Jose Godoy Espinoza"
    set(45, 2, "Nombre: " + nombre_elabora)
    # set(54, 5, "")  # TODO fecha de los protocolos

    #    set(50, 12, "Claudio Boris Hurtado G.")
    #   set(51, 12, "Jefe Terreno")
    # set(54,12,"") #TODO fecha de los protocolos
    pass


for i in range(2, max_line):
    to_print = ws1.cell(i, 12).value
    if to_print != 1:
        continue
        pass

    flush_base()

    corr: str = get(i, 3).zfill(3)
    fields["Correlativo"].value = corr

    sector = get(i, 4)
    fields["Area"].value = sectores[sector]

    tag = get(i, 2)
    elemento: Elemento = resolve_elemento(tag, sector)
    fields["Elemento"].value = elemento.nombre

    plano: str = resolve_plano(sector)
    fields["Plano"].value = plano

    toggle_boxes(elemento)

    toggle_approval()

    toggle_signatures()

    destiny_fl: str = destiny_pth + "\\" + corr + "-PCE-" + sector + "+" + tag.replace("/", "_") + ".xlsx"
    base_fl.save(destiny_fl)

    pass

pass
