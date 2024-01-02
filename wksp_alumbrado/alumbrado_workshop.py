import os
import re

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class Alumb:
    correlative_counter: (int, None) = None

    def __init__(self):
        self.tag: (str, None) = None
        self.patio: (str, None) = None
        self.ubicacion: (str, None) = None
        self.tipo: (str, None) = None
        self.corr: int = -1
        pass

    pass


class AlumbradoWorkshop:
    project_pth: str = os.getcwd()
    res_pth: str = project_pth + "\\res"
    origin_pth :str = project_pth + "\\origin_files"
    destiny_pth: str = project_pth + "\\destiny_files"
    src_pth = res_pth + "\\matriz.txt"

    starting_corr = 1

    patios_excluidos = [
        # "CI",
        # "PK",
        # "PJ",
        # "PZ",
        # "PATR",
    ]

    data: list[Alumb] = []

    @classmethod
    def fetch(cls):
        Alumb.correlative_counter = cls.starting_corr
        source = open(cls.src_pth)
        for line in source:
            pts = line.split("\t")
            al = Alumb()
            al.tag = pts[0].strip()
            al.patio = pts[1].strip()
            al.tipo = pts[2].strip()
            try:
                al.ubicacion = pts[3].strip()
            except IndexError:
                al.ubicacion = ""

            if al.patio in AlumbradoWorkshop.patios_excluidos:
                continue

            al.corr = str(Alumb.correlative_counter).zfill(3)
            Alumb.correlative_counter += 1
            cls.data.append(al)
            pass
        source.close()
        pass

    @classmethod
    def resolve_patio(cls, patio):
        if patio == "PK":
            return "Patio 500kV"
            pass
        elif patio == "PJ":
            return "Patio 220kV"
            pass
        elif patio == "PATR":
            return "Patio ATR"
            pass
        elif patio == "PZ":
            return "Patio Reactores"
            pass
        elif patio == "CI":
            return "Caminos Interiores"
        pass

    @classmethod
    def resolve_sector(cls, sector):
        if sector == "":
            return ""
        return re.sub("D[KJ]", "Diagonal ", sector)
        pass

    @classmethod
    def ensure_paths(cls):
        try:
            os.makedirs(cls.res_pth, exist_ok=True)
            os.makedirs(cls.origin_pth, exist_ok=True)
            os.makedirs(cls.destiny_pth, exist_ok=True)
        except:
            pass
        pass

    pass


class AlumbradoCanalizadoWorkshop:
    # todo

    @classmethod
    def fetch(cls):
        #todo datos por tipo alumb
        pass

    @classmethod
    def gen(cls):
        for d in AlumbradoWorkshop.data:
            cls.inprint(d)
            nm = str(d.corr) + "-PC-" + d.patio + "-" + d.tag.replace("-", "") + ".xlsx"
            pth = AlumbradoWorkshop.destiny_pth + "\\" + nm
            cls.working_template.save(pth)
            pass

    @classmethod
    def inprint(cls, data):

        pass


class AlumbradoMontajeWorkshop:
    src_pth = AlumbradoWorkshop.res_pth + "\\matriz.txt"

    working_template: Workbook = openpyxl.load_workbook(
        AlumbradoWorkshop.res_pth + "\\Protocolo de montaje de luminarias  SSEE Parinas.xlsx")

    ws: Worksheet = working_template.worksheets[0]

    fields: dict[str:Cell] = {
        "correlativo": ws.cell(5, 12),
        "fecha_trabajo": ws.cell(6, 12),
        "descripcion": ws.cell(7, 4),
        "area_trabajo": ws.cell(7, 8),
        "elaboro_nombre": ws.cell(36, 2),
        "elaboro_cargo": ws.cell(38, 2),
        "elaboro_fecha": ws.cell(40, 2),
        "revisa_nombre": ws.cell(36, 5),
        "revisa_cargo": ws.cell(38, 5),
        "reviso_fecha": ws.cell(40, 5)
    }

    @classmethod
    def fetch(cls):
        pass

    @classmethod
    def gen(cls):
        for d in AlumbradoWorkshop.data:
            cls.inprint(d)
            nm = str(d.corr) + "-PML-" + d.patio + "-" + d.tag.replace("-", "") + ".xlsx"
            pth = AlumbradoWorkshop.destiny_pth + "\\" + nm
            cls.working_template.save(pth)
            pass

        pass

    @classmethod
    def inprint(cls, alumb: Alumb):
        cls.fields["correlativo"].value = alumb.corr
        cls.fields["fecha_trabajo"].value = "23-10-2023"
        cls.fields["descripcion"].value = alumb.tipo + " / " + alumb.tag
        cls.fields["area_trabajo"].value = cls.translate_area(alumb.patio, alumb.ubicacion)
        cls.fields["elaboro_nombre"].value = "Nombre: Camilo Miño Miño"
        cls.fields["elaboro_cargo"].value = "Cargo: Supervisor Eléctrico"
        cls.fields["revisa_nombre"].value = "Nombre: Cheslau Mankowski"
        cls.fields["revisa_cargo"].value = "Cargo: Administrador de Contrato"
        pass

    @classmethod
    def translate_area(cls, patio, sector):
        val = AlumbradoWorkshop.resolve_patio(patio)
        val += " " + AlumbradoWorkshop.resolve_sector(sector)

        return val

    pass


class AlumbradoTendidoWorkshop:
    # todo
    pass


class AlumbradoConexionadoWorkshop:
    # todo
    pass


kinds = [
    "Montaje",
    #"Canalizado",
    #"Tendido",
    #"Conexionado",
]


def start(kinds: list[str]):
    AlumbradoWorkshop.ensure_paths()
    AlumbradoWorkshop.fetch()
    if "Montaje" in kinds:
        AlumbradoMontajeWorkshop.fetch()
        AlumbradoMontajeWorkshop.gen()
    if "Canalizado" in kinds:
        AlumbradoCanalizadoWorkshop.fetch()
        AlumbradoCanalizadoWorkshop.gen()
    if "Tendido" in kinds:
        AlumbradoTendidoWorkshop.fetch()
        AlumbradoTendidoWorkshop.gen()
    if "Conexionado" in kinds:
        AlumbradoConexionadoWorkshop.fetch()
        AlumbradoConexionadoWorkshop.gen()
    pass


start(kinds)
pass
