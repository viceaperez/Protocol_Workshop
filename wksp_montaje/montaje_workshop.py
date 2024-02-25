import os

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

project_pth: str = os.getcwd()
res_pth: str = project_pth + "\\res"
destiny_pth: str = project_pth + "\\destiny_files"
base_fl_pth: str = res_pth + "\\SNN4008-E-MMI-01-ELE-002 Montaje de Gabinetes REV.0.xlsx"
base_fl: Workbook = openpyxl.load_workbook(base_fl_pth)
lista_fl_pth: str = res_pth + "\\LOG Montaje de Gabienetes MMI.xlsx"
lista_fl: Workbook = openpyxl.load_workbook(lista_fl_pth, read_only=True)


def ensure_paths():
    lst: list[str] = ["res", "destiny_files"]
    dir_tree: list[str] = os.listdir(project_pth)
    for pth in lst:
        if pth not in dir_tree:
            os.mkdir(pth)
        pass
    pass


ensure_paths()

sectores = {
    "sjd3": "Sala 3",
    "sjd4": "Sala 4",
    "sjd5": "Sala 5",
    "sjd6": "Sala 6",
    "ssgg": "Sala Servicios Generales",
}

planos = {
    "sjd3": "SNN4008-E-MMI-12-EL-PL-0002-L0001",
    "sjd4": "SNN4008-E-MMI-12-EL-PL-0002-L0001",
    "sjd5": "SNN4008-E-MMI-12-EL-PL-0003-L0001",
    "sjd6": "SNN4008-E-MMI-12-EL-PL-0003-L0001",
    "ssgg": "SNN4008-E-MMI-12-EL-PL-0001-L0001",
}

ws: Worksheet = base_fl.worksheets[0]
ws1: Worksheet = lista_fl.worksheets[0]


def set(r, c, val):
    ws.cell(r, c).value = val
    pass


def get(r, c):
    val = str(ws1.cell(r, c).value)
    if val == "None":
        return "-"
    if val == "0":
        return "-"
    return val


max_line = ws1.max_row


def resolve_plano(sector):
    return planos[sector]
    pass

def toggle_approval():
    for j in range(0, 6):
        set(13 + j, 13, "✔")
    pass


def toggle_signatures():
    # TODO: ACTUALIZAR NOMBE SUPERVISOR
    set(30, 3, "Jose Godoy Espinoza")
    set(34, 3, "Supervisor Eléctrico")

    set(30, 8, "Claudio Boris Hurtado G.")
    set(34, 8, "Jefe Terreno")
    pass


for i in range(2, max_line):
    to_print = ws1.cell(i, 12).value
    if to_print != 1:
        continue
        pass

    corr: str = get(i, 3).zfill(3)
    set(7, 19, "Correlativo: " + corr)

    sector = get(i, 4)
    set(8, 5, sectores[sector])
    tag = get(i, 2)
    set(7, 5, tag)

    plano: str = resolve_plano(sector)
    set(9, 6, plano)

    fecha: str = "24/02/2024"
    set(7, 14, fecha)
    set(35, 3, fecha)
    set(35, 8, fecha)

    toggle_approval()

    toggle_signatures()

    destiny_fl: str = destiny_pth + "\\" + corr + "-PMT-" + tag.replace("/", "_") + ".xlsx"
    base_fl.save(destiny_fl)

    pass

pass
