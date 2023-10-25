import os

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

project_pth: str = os.getcwd()
res_pth: str = project_pth + "\\res"
destiny_pth: str = project_pth + "\\destiny_files"
origin_pth: str = project_pth + "\\origin_files"

fl = open(project_pth + "\\2fetch.txt","r")
out = open(destiny_pth+"\\out.txt","w")

wb:Workbook = openpyxl.load_workbook(res_pth+"\\lista.xlsx",read_only=True)
ws:Worksheet = wb.worksheets[0]

arr = []
for ws_idx in range(1, ws.max_row + 1):
    arr.append([ws.cell(ws_idx,3).value,ws.cell(ws_idx, 2).value])
    pass
pass



for line in fl:
    for idx in range(len(arr)):
        if line.strip() == arr[idx][0]:
            out.write(arr[idx][0]+"\tConstruccion de Malla Puesta a Tierra\t"+arr[idx][1]+"\n")
            pass
        pass
    pass
out.close()


