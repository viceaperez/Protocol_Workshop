import os
import re
from enum import Enum

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

# TODO: MODIFICAR EL ORIGEN DE DATOS DE ELEMENTOS PARA DISTINTOS PATIOS
origen = "elementosPZ.xlsx"

project_pth: str = os.getcwd()
res_pth: str = project_pth + "\\res"


class Cable:
    def __init__(self):
        self.calibre: str = "-"
        self.aislacion: str = "-"
        self.equipos: (str, float) = "-"
        self.estructura: (str, float) = "-"
        self.tierra: (str, float) = "-"
        pass

    pass


class Elemento:

    def __init__(self):
        self.tipo = None
        self.cables: list[Cable] = []
        self.soldaduras: list[int] = []
        self.categorias: list = []
        self.nombre = ""
        self.patron: re.Pattern

    pass


class Categoria(Enum):
    BANCODUCTO = "Bancoductos"
    ESTRUCTURAL = "Estructurales"
    ESCALERILLA = "Escalerillas"
    CAMARA = "Camaras"
    EQUIPO = "Equipos"
    TRANSFORMADOR_OTROS = "totros"
    pass


elementos: dict[str:Elemento] = {}
''''''
wb: Workbook = openpyxl.load_workbook(res_pth + "\\" + origen)
for ws in wb:
    ws: Worksheet
    el = Elemento()
    lk = False
    el.nombre = ws.cell(22, 2).value
    el.id = ws.title
    el.patron = re.compile(ws.cell(21, 2).value, re.IGNORECASE)
    for i in range(1, 13):
        el.soldaduras.append(ws.cell(i, 2).value)
        pass
    for i in range(15, 20):
        cb = Cable()
        cb.calibre = ws.cell(i, 1).value
        cb.aislacion = ws.cell(i, 2).value
        cb.equipos = ws.cell(i, 3).value
        cb.estructura = ws.cell(i, 4).value
        cb.tierra = ws.cell(i, 5).value
        el.cables.append(cb)
        pass
    for i in range(1, 7):
        if ws.cell(i, 5).value:
            for cat in Categoria:
                if cat.name == ws.cell(i, 4).value:
                    el.categorias.append(cat)
                    break
                    pass
                pass
            pass
        pass
        pass
    elementos[el.nombre] = el
    pass

''''''
