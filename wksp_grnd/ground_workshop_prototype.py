import os

import openpyxl
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from tipos import Elemento, elementos

project_pth: str = os.getcwd()
res_pth: str = project_pth + "\\res"
destiny_pth: str = project_pth + "\\destiny_files"
base_fl_pth: str = res_pth + "\\base.xlsx"
base_fl: Workbook = openpyxl.load_workbook(base_fl_pth)
lista_fl_pth: str = res_pth + "\\lista.xlsx"
lista_fl: Workbook = openpyxl.load_workbook(lista_fl_pth, read_only=True)


def ensure_paths():
    lst: list[str] = ["res", "destiny_files"]
    dir_tree: list[str] = os.listdir(project_pth)
    for pth in lst:
        if pth not in dir_tree:
            os.mkdir(pth)
        pass
    pass


ensure_paths()

sectores = {
    "PK": "Patio 500kV",
    "PJ": "Patio 220kV",
    "PATR": "Patio ATR",
    "PZ": "Patio Reactores"
}

planos = {
    "planta": "SNN4008-E-PRN-13-EL-PL-0001-L0001",
    "pk": "SNN4008-E-PRN-13-EL-PL-0001-L0002",
    "pk-d": "SNN4008-E-PRN-13-EL-PL-0001-L0003",
    "pj": "SNN4008-E-PRN-13-EL-PL-0006-L0001",
    "pj-d": "SNN4008-E-PRN-13-EL-PL-0006-L0002",
    "atr_r": "SNN4008-E-PRN-13-EL-PL-0007-L0001",
    "atr_r-d": "SNN4008-E-PRN-13-EL-PL-0007-L0002",
}

ws: Worksheet = base_fl.worksheets[0]
ws1: Worksheet = lista_fl.worksheets[0]

checks: dict[str:Cell] = {
    "Bancoductos": ws.cell(12, 12),
    "Estructurales": ws.cell(12, 18),
    "Escalerillas": ws.cell(12, 24),
    "Camaras": ws.cell(14, 12),
    "Equipos": ws.cell(14, 18),
    "totros": ws.cell(14, 24),
}

solds: list[Cell] = []
for i in range(0, 6):
    for j in range(0, 2):
        solds.append(ws.cell(18 + i, 22 + (j * 7)))
    pass


def set(r, c, val):
    ws.cell(r, c).value = val
    pass


def get(r, c):
    val = str(ws1.cell(r, c).value)
    if val == "None":
        return "-"
    if val == "0":
        return "-"
    return val


max_line = ws1.max_row


def flush_base():
    for check in checks.values():
        check.value = ""
        pass

    for j in range(0, 5):
        set(19 + j, 3, "-")
        set(19 + j, 7, "-")
        set(19 + j, 9, "-")
        set(19 + j, 11, "-")
        set(19 + j, 13, "-")

    for j in range(0, 6):
        for k in range(0, 2):
            set(18 + j, 22 + (k * 7), "-")
            pass
        pass
    pass


def resolve_elemento(tag_elem) -> Elemento:
    for tipo in elementos.values():
        if tipo.patron.match(tag_elem):
            return tipo
    pass


def resolve_plano(sector):
    if sector == "PK":
        return planos["pk"]
    if sector == "PJ":
        return planos["pj"]
    if sector == "PATR" or sector == "PZ":
        return planos["atr_r"]
    pass


def toggle_check(elemento: Elemento):
    for ck in checks.keys():
        for cat in elemento.categorias:
            if ck.lower() == cat.value.lower():
                checks[ck].value = "✔"
                pass
    pass


def toggle_cables(elemento: Elemento):
    for i in range(0, len(elemento.cables)):
        set(19 + i, 3, elemento.cables[i].calibre)
        set(19 + i, 7, elemento.cables[i].aislacion)
        set(19 + i, 9, elemento.cables[i].equipos)
        set(19 + i, 11, elemento.cables[i].estructura)
        set(19 + i, 13, elemento.cables[i].tierra)
        set(19 + i, 15, "=SUM(I" + str(19 + i) + ":N" + str(19 + i) + ")")
    pass


def toggle_soldaduras(elemento: Elemento):
    for i in range(len(solds)):
        solds[i].value = elemento.soldaduras[i]
        pass


def toggle_approval():
    for j in range(0, 12):
        set(25 + j, 15, "✔")
    pass


def toggle_signatures():
    # TODO: ACTUALIZAR NOMBE SUPERVISOR
    set(50, 5, "Jose Godoy Espinoza")
    set(51, 5, "Supervisor Eléctrico")
    # set(54, 5, "")  # TODO fecha de los protocolos

    set(50, 12, "Claudio Boris Hurtado G.")
    set(51, 12, "Jefe Terreno")
    # set(54,12,"") #TODO fecha de los protocolos
    pass


for i in range(2, max_line):
    to_print = ws1.cell(i, 12).value
    if to_print != 1:
        continue
        pass

    flush_base()

    corr: str = get(i, 3).zfill(3)
    set(6, 29, corr)

    sector = get(i, 4)
    set(8, 5, sectores[sector])
    tag = get(i, 2)
    set(8, 24, tag)

    elemento: Elemento = resolve_elemento(tag)
    set(7, 6, elemento.nombre)

    plano: str = resolve_plano(sector)
    set(9, 9, plano)

    toggle_check(elemento)

    toggle_cables(elemento)

    toggle_soldaduras(elemento)

    toggle_approval()

    toggle_signatures()

    destiny_fl: str = destiny_pth + "\\" + corr + "-PMT-" + tag.replace("/", "_") + ".xlsx"
    base_fl.save(destiny_fl)

    pass

pass
