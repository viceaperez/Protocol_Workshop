import os

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

project_pth: str = os.getcwd()
res_pth: str = project_pth + "\\res"
destiny_pth: str = project_pth + "\\destiny_files"
origin_pth: str = project_pth + "\\origin_files"

lst = os.listdir(origin_pth)

lst_log: dict[str:str] = {}

wb: Workbook = openpyxl.load_workbook(res_pth + "\\lista.xlsx")
ws: Worksheet = wb.worksheets[0]
for i in range(1, ws.max_row + 1):
    lst_log[ws.cell(i, 2).value] = ws.cell(i, 3).value
    pass

for e in lst:
    tmp: Workbook = openpyxl.load_workbook(origin_pth + "\\" + e)
    sh: Worksheet = tmp.worksheets[0]
    crr = sh.cell(5, 29).value
    tag = sh.cell(8, 26).value

    if crr is None:
        crr = "___"
        if tag in lst_log:
            crr = lst_log[tag]
        pass
    tmp.save(destiny_pth + "\\" + crr + "-PMT-" + tag.split("+")[1].replace("/", "_") + ".xlsx")
